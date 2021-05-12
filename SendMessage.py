import openpyxl

from whatsapp import WhatsApp

whatsapp = WhatsApp(100, session="mysession")
wb = openpyxl.load_workbook("UserInput/ListOfGroups.xlsx")
sheet = wb.active
for row in range(sheet.max_row-1):
    f = open('UserInput/SendMessage.txt')
    message = f.read();
    whatsapp.send_message(str(sheet.cell(row+2,1).value),message)
    row=row+1
    f.close()
wb.close()
whatsapp.quit()