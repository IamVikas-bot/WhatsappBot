import openpyxl

from whatsapp import WhatsApp

whatsapp = WhatsApp(100, session="mysession")
wb = openpyxl.load_workbook("UserInput/ListOfGroups.xlsx")
sheet = wb.active
wb1 = openpyxl.load_workbook("UserInput/GroupLinks.xlsx")
sheet1 = wb1.active
new_groups = []
row=0;
for row in range(sheet1.max_row-1):
    name = whatsapp.join_group(str(sheet1.cell(row+2,1).value))
    new_groups.append(name)
    row=row+1
wb1.close()

for name in new_groups:
   if not name:
    print("Blank Value")
   else:
    sheet.append(name)

wb.close()
wb.save('ListOfGroups.xlsx')
whatsapp.quit()