import os
import openpyxl

from whatsapp import WhatsApp

whatsapp = WhatsApp(100, session="mysession")
wb = openpyxl.load_workbook("UserInput/ListOfGroups.xlsx")
sheet = wb.active
for row in range(sheet.max_row-1):
    whatsapp.send_picture(str(sheet.cell(row+2,1).value),os.getcwd()+'\ClaimsData.png')
    row=row+1

whatsapp.quit()